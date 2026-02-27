import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Colores ──────────────────────────────────────────────────────────────────
COLOR_NEGRO  = "000000"
COLOR_MORADO = "8563E1"
COLOR_LILA   = "E8E0FA"
COLOR_BLANCO = "FFFFFF"

st.set_page_config(page_title="Verificador de Áreas", page_icon="🏠", layout="centered")

st.markdown("""
<style>
.titulo-negro  {background:#000;color:#fff;padding:10px 16px;border-radius:6px;font-weight:bold;font-size:17px;margin-bottom:8px;}
.titulo-morado {background:#8563E1;color:#fff;padding:8px 16px;border-radius:6px;font-weight:bold;font-size:14px;margin:6px 0;}
.resultado-ok  {background:#d4edda;color:#155724;padding:12px;border-radius:6px;font-weight:bold;margin-top:6px;}
.resultado-mal {background:#f8d7da;color:#721c24;padding:12px;border-radius:6px;font-weight:bold;margin-top:6px;}
.resultado-info{background:#d1ecf1;color:#0c5460;padding:12px;border-radius:6px;font-weight:bold;margin-top:6px;}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="titulo-negro">🏠 VERIFICADOR DE ÁREAS — INSPECCIÓN DE DEPARTAMENTOS</div>', unsafe_allow_html=True)

# ── Datos generales ──────────────────────────────────────────────────────────
st.subheader("Datos del Departamento")
c1, c2 = st.columns(2)
with c1:
    proyecto     = st.text_input("Nombre del Proyecto / Inmobiliaria")
    departamento = st.text_input("N° de Departamento")
with c2:
    inspector = st.text_input("Inspector")
    fecha     = st.date_input("Fecha de Inspección")

area_ofrecida = st.number_input("Área ofrecida por la inmobiliaria — Departamento (m²)", min_value=0.0, step=0.01, format="%.2f")
factor_ajuste = st.number_input("Factor de ajuste principal (muros, placas, columnas)", min_value=1.0, value=1.13, step=0.01, format="%.2f")

TOLERANCIA_PCT = 2.5

# ── Estado inicial ────────────────────────────────────────────────────────────
AMBIENTES_DEFAULT       = ["Dormitorio Principal","Baño Principal","Dormitorio Secundario 1",
                           "Dormitorio Secundario 2","Baño Secundario","Sala - Comedor",
                           "Cocina","Lavandería","Estudio / Escritorio","Balcón / Terraza","Pasillo / Hall"]
COMPLEMENTARIOS_DEFAULT = ["Cochera","Depósito"]

def amb_vacio(nombre):
    return {"nombre": nombre, "medidas": [{"largo": 0.0, "ancho": 0.0}], "area_ofrecida": 0.0}

if "ambientes"       not in st.session_state:
    st.session_state.ambientes       = [{"nombre": n, "medidas": [{"largo": 0.0, "ancho": 0.0}]} for n in AMBIENTES_DEFAULT]
if "complementarios" not in st.session_state:
    st.session_state.complementarios = [amb_vacio(n) for n in COMPLEMENTARIOS_DEFAULT]
if "factor_comp"     not in st.session_state:
    st.session_state.factor_comp     = 1.10

# ── Renderizar ambientes principales (sin área ofrecida) ─────────────────────
def render_ambientes(lista_key):
    lista   = st.session_state[lista_key]
    totales = []
    for i, ambiente in enumerate(lista):
        with st.expander(f"📐 {ambiente['nombre']}", expanded=True):
            lista[i]["nombre"] = st.text_input(
                "Nombre del ambiente", value=ambiente["nombre"], key=f"{lista_key}_nom_{i}"
            )
            subtotal = 0.0
            for j, medida in enumerate(ambiente["medidas"]):
                ca, cb, cc = st.columns(3)
                with ca:
                    largo = st.number_input("Largo (m)", min_value=0.0, step=0.01,
                                            value=float(medida["largo"]),
                                            key=f"{lista_key}_l_{i}_{j}", format="%.2f")
                with cb:
                    ancho = st.number_input("Ancho (m)", min_value=0.0, step=0.01,
                                            value=float(medida["ancho"]),
                                            key=f"{lista_key}_a_{i}_{j}", format="%.2f")
                with cc:
                    st.metric(f"Área Parcial {j+1} (m²)", f"{largo*ancho:.2f}")
                lista[i]["medidas"][j] = {"largo": largo, "ancho": ancho}
                subtotal += largo * ancho

            bca, bcb = st.columns(2)
            with bca:
                if len(ambiente["medidas"]) < 5:
                    if st.button("➕ Agregar medida", key=f"{lista_key}_add_m_{i}"):
                        lista[i]["medidas"].append({"largo": 0.0, "ancho": 0.0})
                        st.rerun()
            with bcb:
                if len(ambiente["medidas"]) > 1:
                    if st.button("➖ Quitar última medida", key=f"{lista_key}_del_m_{i}"):
                        lista[i]["medidas"].pop()
                        st.rerun()

            st.markdown(f'<div style="color:#8563E1;font-weight:bold;">Subtotal {lista[i]["nombre"]}: {subtotal:.2f} m²</div>',
                        unsafe_allow_html=True)
            totales.append(subtotal)

    baa, bab = st.columns(2)
    with baa:
        if st.button("➕ Agregar ambiente", key=f"{lista_key}_add_amb"):
            lista.append({"nombre": "Nuevo Ambiente", "medidas": [{"largo": 0.0, "ancho": 0.0}]})
            st.rerun()
    with bab:
        if len(lista) > 1:
            if st.button("🗑️ Eliminar último ambiente", key=f"{lista_key}_del_amb"):
                lista.pop()
                st.rerun()
    return totales

# ── Renderizar complementarios (con área ofrecida individual) ─────────────────
def render_complementarios():
    lista   = st.session_state.complementarios
    totales = []
    for i, ambiente in enumerate(lista):
        with st.expander(f"📐 {ambiente['nombre']}", expanded=True):
            lista[i]["nombre"] = st.text_input(
                "Nombre", value=ambiente["nombre"], key=f"comp_nom_{i}"
            )
            # Área ofrecida individual
            lista[i]["area_ofrecida"] = st.number_input(
                f"Área ofrecida por inmobiliaria — {lista[i]['nombre']} (m²)",
                min_value=0.0, step=0.01,
                value=float(ambiente.get("area_ofrecida", 0.0)),
                key=f"comp_ofrecida_{i}", format="%.2f"
            )

            subtotal = 0.0
            for j, medida in enumerate(ambiente["medidas"]):
                ca, cb, cc = st.columns(3)
                with ca:
                    largo = st.number_input("Largo (m)", min_value=0.0, step=0.01,
                                            value=float(medida["largo"]),
                                            key=f"comp_l_{i}_{j}", format="%.2f")
                with cb:
                    ancho = st.number_input("Ancho (m)", min_value=0.0, step=0.01,
                                            value=float(medida["ancho"]),
                                            key=f"comp_a_{i}_{j}", format="%.2f")
                with cc:
                    st.metric(f"Área Parcial {j+1} (m²)", f"{largo*ancho:.2f}")
                lista[i]["medidas"][j] = {"largo": largo, "ancho": ancho}
                subtotal += largo * ancho

            bca, bcb = st.columns(2)
            with bca:
                if len(ambiente["medidas"]) < 5:
                    if st.button("➕ Agregar medida", key=f"comp_add_m_{i}"):
                        lista[i]["medidas"].append({"largo": 0.0, "ancho": 0.0})
                        st.rerun()
            with bcb:
                if len(ambiente["medidas"]) > 1:
                    if st.button("➖ Quitar última medida", key=f"comp_del_m_{i}"):
                        lista[i]["medidas"].pop()
                        st.rerun()

            area_real_comp = subtotal * st.session_state.factor_comp
            st.markdown(f'<div style="color:#8563E1;font-weight:bold;">Subtotal {lista[i]["nombre"]}: {subtotal:.2f} m² → Área real: {area_real_comp:.2f} m²</div>',
                        unsafe_allow_html=True)

            # Verificación inmediata por ambiente complementario
            ao = lista[i].get("area_ofrecida", 0.0)
            if ao > 0:
                tol     = ao * (TOLERANCIA_PCT / 100)
                diff    = area_real_comp - ao
                pct     = (diff / ao) * 100
                if abs(diff) <= tol:
                    st.markdown(f'<div class="resultado-ok">✅ CONFORME — Diferencia {diff:+.2f} m² ({pct:+.1f}%) dentro de ±{TOLERANCIA_PCT}%</div>', unsafe_allow_html=True)
                elif diff < 0:
                    st.markdown(f'<div class="resultado-mal">❌ NO CONFORME — MENOR en {abs(diff):.2f} m² ({abs(pct):.1f}%)</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="resultado-info">ℹ️ MAYOR en {diff:.2f} m² ({pct:.1f}%)</div>', unsafe_allow_html=True)

            totales.append(subtotal)

    baa, bab = st.columns(2)
    with baa:
        if st.button("➕ Agregar área complementaria", key="comp_add_amb"):
            lista.append(amb_vacio("Nueva Área"))
            st.rerun()
    with bab:
        if len(lista) > 1:
            if st.button("🗑️ Eliminar última", key="comp_del_amb"):
                lista.pop()
                st.rerun()
    return totales

# ── SECCIÓN PRINCIPAL ─────────────────────────────────────────────────────────
st.markdown('<div class="titulo-negro">REGISTRO DE MEDIDAS DE AMBIENTES - DEPARTAMENTO</div>', unsafe_allow_html=True)
totales_principal   = render_ambientes("ambientes")
total_cc_principal  = sum(totales_principal)
area_real_principal = total_cc_principal * factor_ajuste

st.markdown('<div class="titulo-morado">RESUMEN — ÁREA PRINCIPAL</div>', unsafe_allow_html=True)
r1, r2 = st.columns(2)
r1.metric("Total Cara a Cara (m²)",    f"{total_cc_principal:.2f}")
r2.metric("Área Real con Factor (m²)", f"{area_real_principal:.2f}")

# ── SECCIÓN COMPLEMENTARIA ────────────────────────────────────────────────────
st.markdown('<div class="titulo-negro">ÁREAS COMPLEMENTARIAS (no incluidas en área del departamento)</div>', unsafe_allow_html=True)
factor_comp = st.number_input("Factor de ajuste complementarios", min_value=1.0,
                               value=float(st.session_state.factor_comp), step=0.01, format="%.2f")
st.session_state.factor_comp = factor_comp

totales_comp   = render_complementarios()
total_cc_comp  = sum(totales_comp)
area_real_comp = total_cc_comp * factor_comp

st.markdown('<div class="titulo-morado">RESUMEN — ÁREAS COMPLEMENTARIAS</div>', unsafe_allow_html=True)
rc1, rc2 = st.columns(2)
rc1.metric("Total Cara a Cara (m²)",  f"{total_cc_comp:.2f}")
rc2.metric("Área Real Total (m²)",    f"{area_real_comp:.2f}")

# ── RESULTADO FINAL DEPARTAMENTO ──────────────────────────────────────────────
st.divider()
st.markdown('<div class="titulo-negro">📊 RESULTADO FINAL DE INSPECCIÓN</div>', unsafe_allow_html=True)

# Departamento
diferencia = area_real_principal - area_ofrecida
st.markdown('<div class="titulo-morado">Departamento</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)
col1.metric("Área Real (m²)",     f"{area_real_principal:.2f}")
col2.metric("Área Ofrecida (m²)", f"{area_ofrecida:.2f}")
col3.metric("Diferencia (m²)",    f"{diferencia:+.2f}")

if area_ofrecida > 0:
    tol_depto = area_ofrecida * (TOLERANCIA_PCT / 100)
    pct_depto = (diferencia / area_ofrecida) * 100
    if abs(diferencia) <= tol_depto:
        st.markdown(f'<div class="resultado-ok">✅ CONFORME — Diferencia {diferencia:+.2f} m² ({pct_depto:+.1f}%) dentro de tolerancia ±{TOLERANCIA_PCT}% ({tol_depto:.2f} m²).</div>', unsafe_allow_html=True)
    elif diferencia < 0:
        st.markdown(f'<div class="resultado-mal">❌ NO CONFORME — Área real MENOR en {abs(diferencia):.2f} m² ({abs(pct_depto):.1f}%). Supera tolerancia ±{TOLERANCIA_PCT}%.</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div class="resultado-info">ℹ️ Área real MAYOR en {diferencia:.2f} m² ({pct_depto:.1f}%). Supera tolerancia ±{TOLERANCIA_PCT}%.</div>', unsafe_allow_html=True)

# Complementarios individuales
st.markdown('<div class="titulo-morado">Áreas Complementarias</div>', unsafe_allow_html=True)
for amb in st.session_state.complementarios:
    ao_c      = amb.get("area_ofrecida", 0.0)
    subtotal_c = sum(m["largo"] * m["ancho"] for m in amb["medidas"])
    ar_c      = subtotal_c * factor_comp
    if ao_c > 0 and ar_c > 0:
        diff_c = ar_c - ao_c
        tol_c  = ao_c * (TOLERANCIA_PCT / 100)
        pct_c  = (diff_c / ao_c) * 100
        c1, c2, c3 = st.columns(3)
        c1.metric(f"{amb['nombre']} — Real (m²)",     f"{ar_c:.2f}")
        c2.metric(f"{amb['nombre']} — Ofrecida (m²)", f"{ao_c:.2f}")
        c3.metric("Diferencia (m²)",                  f"{diff_c:+.2f}")
        if abs(diff_c) <= tol_c:
            st.markdown(f'<div class="resultado-ok">✅ {amb["nombre"]} CONFORME — Diferencia {diff_c:+.2f} m² ({pct_c:+.1f}%) dentro de ±{TOLERANCIA_PCT}%.</div>', unsafe_allow_html=True)
        elif diff_c < 0:
            st.markdown(f'<div class="resultado-mal">❌ {amb["nombre"]} NO CONFORME — MENOR en {abs(diff_c):.2f} m² ({abs(pct_c):.1f}%). Supera ±{TOLERANCIA_PCT}%.</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="resultado-info">ℹ️ {amb["nombre"]} MAYOR en {diff_c:.2f} m² ({pct_c:.1f}%). Supera ±{TOLERANCIA_PCT}%.</div>', unsafe_allow_html=True)

# ── EXPORTAR EXCEL ────────────────────────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def cell_style(ws, row, col, value, bg, fg, bold=False, align="left", number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = hex_fill(bg)
    cell.font      = Font(color=fg, bold=bold, name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_format:
        cell.number_format = number_format
    return cell

def titulo_negro(ws, row, texto, size=12):
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.fill      = PatternFill("solid", fgColor=COLOR_NEGRO)
    c.font      = Font(color=COLOR_BLANCO, bold=True, name="Arial", size=size)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

def amb_tiene_medidas(amb):
    return any(m["largo"] > 0 and m["ancho"] > 0 for m in amb["medidas"])

def escribir_ambientes_excel(ws, lista, row, con_area_ofrecida=False):
    subtotal_rows = []
    for amb in lista:
        if not amb_tiene_medidas(amb):
            continue

        ws.merge_cells(f"A{row}:E{row}")
        cell_style(ws, row, 1, amb["nombre"], COLOR_LILA, COLOR_MORADO, bold=True)
        ws.row_dimensions[row].height = 16
        row += 1

        if con_area_ofrecida:
            ao = amb.get("area_ofrecida", 0.0)
            cell_style(ws, row, 1, "   Área ofrecida por inmobiliaria (m²):", COLOR_BLANCO, COLOR_NEGRO)
            cell_style(ws, row, 2, ao, COLOR_LILA, COLOR_MORADO, number_format="0.00")
            ws.row_dimensions[row].height = 15
            row += 1

        area_parcial_rows = []
        for j, m in enumerate(amb["medidas"]):
            if m["largo"] == 0 and m["ancho"] == 0:
                continue
            parcial = m["largo"] * m["ancho"]
            cell_style(ws, row, 1, f"   Medida {j+1}", COLOR_BLANCO, COLOR_NEGRO)
            cell_style(ws, row, 2, m["largo"],          COLOR_BLANCO, COLOR_NEGRO, number_format="0.00")
            cell_style(ws, row, 3, m["ancho"],          COLOR_BLANCO, COLOR_NEGRO, number_format="0.00")
            cell_style(ws, row, 4, parcial,             COLOR_BLANCO, COLOR_NEGRO, number_format="0.00")
            area_parcial_rows.append(row)
            ws.row_dimensions[row].height = 15
            row += 1

        if area_parcial_rows:
            subtotal_formula = f"=SUM(D{area_parcial_rows[0]}:D{area_parcial_rows[-1]})"
            cell_style(ws, row, 1, f"   Subtotal {amb['nombre']}", COLOR_BLANCO, COLOR_MORADO, bold=True)
            sc = ws.cell(row=row, column=5, value=subtotal_formula)
            sc.fill          = hex_fill(COLOR_BLANCO)
            sc.font          = Font(color=COLOR_MORADO, bold=True, name="Arial", size=10)
            sc.number_format = "0.00"
            subtotal_rows.append((row, amb))
            ws.row_dimensions[row].height = 15
            row += 1

    return row, subtotal_rows

def resultado_color(diff, tol):
    if abs(diff) <= tol:
        return "d4edda", "155724", "CONFORME"
    elif diff < 0:
        return "f8d7da", "721c24", "NO CONFORME"
    else:
        return "d1ecf1", "0c5460", "OBSERVACIÓN"

def exportar_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "REGISTRO"

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22

    row = 1
    titulo_negro(ws, row, "REGISTRO DE MEDIDAS DE AMBIENTES - DEPARTAMENTO", size=12)
    row += 1

    info = [
        ("Proyecto / Inmobiliaria:", proyecto),
        ("N° Departamento:",         departamento),
        ("Inspector:",               inspector),
        ("Fecha:",                   str(fecha)),
        ("Área ofrecida departamento (m²):", area_ofrecida),
        ("Factor de ajuste:",        factor_ajuste),
    ]
    for label, val in info:
        cell_style(ws, row, 1, label, COLOR_BLANCO, COLOR_NEGRO, bold=True)
        if isinstance(val, float):
            cell_style(ws, row, 2, val, COLOR_LILA, COLOR_MORADO, number_format="0.00")
        else:
            cell_style(ws, row, 2, val, COLOR_LILA, COLOR_MORADO)
        ws.row_dimensions[row].height = 15
        row += 1
    row += 1

    headers = ["AMBIENTE / MEDICIÓN", "LARGO (m)", "ANCHO (m)", "ÁREA PARCIAL (m²)", "SUBTOTAL AMBIENTE (m²)"]
    for c, h in enumerate(headers, 1):
        cell_style(ws, row, c, h, COLOR_MORADO, COLOR_BLANCO, bold=True, align="center")
    ws.row_dimensions[row].height = 18
    row += 1

    row, subtotal_data = escribir_ambientes_excel(ws, st.session_state.ambientes, row, con_area_ofrecida=False)
    subtotal_rows = [r for r, _ in subtotal_data]
    row += 1

    total_formula = ("=" + "+".join([f"E{r}" for r in subtotal_rows])) if subtotal_rows else 0
    ws.merge_cells(f"A{row}:D{row}")
    cell_style(ws, row, 1, "TOTAL ÁREA CARA A CARA (m²)", COLOR_NEGRO, COLOR_BLANCO, bold=True)
    tc = ws.cell(row=row, column=5, value=total_formula)
    tc.fill = PatternFill("solid", fgColor=COLOR_NEGRO)
    tc.font = Font(color=COLOR_BLANCO, bold=True, name="Arial", size=10)
    tc.number_format = "0.00"
    total_row = row; row += 1

    ws.merge_cells(f"A{row}:C{row}")
    cell_style(ws, row, 1, "FACTOR DE AJUSTE (muros, placas, columnas):", COLOR_BLANCO, COLOR_NEGRO)
    cell_style(ws, row, 4, factor_ajuste, COLOR_LILA, COLOR_MORADO, number_format="0.00")
    factor_row = row; row += 1

    ws.merge_cells(f"A{row}:D{row}")
    cell_style(ws, row, 1, "ÁREA REAL DEL DEPARTAMENTO (m²)", COLOR_MORADO, COLOR_BLANCO, bold=True)
    ar = ws.cell(row=row, column=5, value=f"=E{total_row}*D{factor_row}")
    ar.fill = PatternFill("solid", fgColor=COLOR_MORADO)
    ar.font = Font(color=COLOR_BLANCO, bold=True, name="Arial", size=10)
    ar.number_format = "0.00"
    area_real_row = row; row += 2

    # ── Complementarias ──
    titulo_negro(ws, row, "ÁREAS COMPLEMENTARIAS (no incluidas en área del departamento)", size=11)
    row += 1
    cell_style(ws, row, 1, "Factor de ajuste complementarios:", COLOR_BLANCO, COLOR_NEGRO)
    cell_style(ws, row, 2, factor_comp, COLOR_LILA, COLOR_MORADO, number_format="0.00")
    ws.row_dimensions[row].height = 15; row += 1; row += 1  # espacio extra

    for c, h in enumerate(headers, 1):
        cell_style(ws, row, c, h, COLOR_MORADO, COLOR_BLANCO, bold=True, align="center")
    ws.row_dimensions[row].height = 18
    row += 1

    row, subtotal_comp_data = escribir_ambientes_excel(ws, st.session_state.complementarios, row, con_area_ofrecida=True)
    subtotal_comp_rows = [r for r, _ in subtotal_comp_data]
    row += 1

    total_comp_formula = ("=" + "+".join([f"E{r}" for r in subtotal_comp_rows])) if subtotal_comp_rows else 0
    ws.merge_cells(f"A{row}:D{row}")
    cell_style(ws, row, 1, "TOTAL ÁREAS COMPLEMENTARIAS (m²)", COLOR_NEGRO, COLOR_BLANCO, bold=True)
    tc2 = ws.cell(row=row, column=5, value=total_comp_formula)
    tc2.fill = PatternFill("solid", fgColor=COLOR_NEGRO)
    tc2.font = Font(color=COLOR_BLANCO, bold=True, name="Arial", size=10)
    tc2.number_format = "0.00"
    total_comp_row = row; row += 1

    ws.merge_cells(f"A{row}:C{row}")
    cell_style(ws, row, 1, "FACTOR DE AJUSTE (muros, placas, columnas):", COLOR_BLANCO, COLOR_NEGRO)
    cell_style(ws, row, 4, factor_comp, COLOR_LILA, COLOR_MORADO, number_format="0.00")
    factor_comp_row = row; row += 1

    ws.merge_cells(f"A{row}:D{row}")
    cell_style(ws, row, 1, "ÁREA REAL COMPLEMENTARIOS (m²)", COLOR_MORADO, COLOR_BLANCO, bold=True)
    ar2 = ws.cell(row=row, column=5, value=f"=E{total_comp_row}*D{factor_comp_row}")
    ar2.fill = PatternFill("solid", fgColor=COLOR_MORADO)
    ar2.font = Font(color=COLOR_BLANCO, bold=True, name="Arial", size=10)
    ar2.number_format = "0.00"
    row += 2

    # ── Resultado final ──
    titulo_negro(ws, row, "RESULTADO DE INSPECCIÓN", size=11)
    row += 1

    # Departamento
    cell_style(ws, row, 1, "DEPARTAMENTO", COLOR_MORADO, COLOR_BLANCO, bold=True)
    ws.row_dimensions[row].height = 16; row += 1

    tol_d  = area_ofrecida * (TOLERANCIA_PCT / 100) if area_ofrecida > 0 else 0
    diff_d = area_real_principal - area_ofrecida
    bg_d, fg_d, estado_d = resultado_color(diff_d, tol_d) if area_ofrecida > 0 else (COLOR_BLANCO, COLOR_NEGRO, "—")
    pct_d  = (diff_d / area_ofrecida * 100) if area_ofrecida > 0 else 0

    for label, val, is_num in [
        ("Área ofrecida (m²):",    area_ofrecida,      True),
        ("Área real (m²):",        area_real_principal, True),
        ("Diferencia (m²):",       diff_d,              True),
        (f"Tolerancia ±{TOLERANCIA_PCT}% (m²):", tol_d, True),
        ("Resultado:",             f"{estado_d} — {diff_d:+.2f} m² ({pct_d:+.1f}%)", False),
    ]:
        cell_style(ws, row, 1, label, COLOR_LILA, COLOR_MORADO, bold=True)
        if is_num:
            cell_style(ws, row, 2, val, COLOR_BLANCO, COLOR_NEGRO, number_format="0.00")
        else:
            ws.merge_cells(f"B{row}:E{row}")
            cell_style(ws, row, 2, val, bg_d, fg_d, bold=True)
        ws.row_dimensions[row].height = 15; row += 1
    row += 1

    # Complementarios
    for amb in st.session_state.complementarios:
        if not amb_tiene_medidas(amb):
            continue
        ao_c = amb.get("area_ofrecida", 0.0)
        st_c = sum(m["largo"] * m["ancho"] for m in amb["medidas"])
        ar_c = st_c * factor_comp
        tol_c = ao_c * (TOLERANCIA_PCT / 100) if ao_c > 0 else 0
        diff_c = ar_c - ao_c
        bg_c, fg_c, estado_c = resultado_color(diff_c, tol_c) if ao_c > 0 else (COLOR_BLANCO, COLOR_NEGRO, "—")
        pct_c = (diff_c / ao_c * 100) if ao_c > 0 else 0

        cell_style(ws, row, 1, amb["nombre"].upper(), COLOR_MORADO, COLOR_BLANCO, bold=True)
        ws.row_dimensions[row].height = 16; row += 1

        for label, val, is_num in [
            ("Área ofrecida (m²):", ao_c,     True),
            ("Área real (m²):",     ar_c,     True),
            ("Diferencia (m²):",    diff_c,   True),
            (f"Tolerancia ±{TOLERANCIA_PCT}% (m²):", tol_c, True),
            ("Resultado:", f"{estado_c} — {diff_c:+.2f} m² ({pct_c:+.1f}%)", False),
        ]:
            cell_style(ws, row, 1, label, COLOR_LILA, COLOR_MORADO, bold=True)
            if is_num:
                cell_style(ws, row, 2, val, COLOR_BLANCO, COLOR_NEGRO, number_format="0.00")
            else:
                ws.merge_cells(f"B{row}:E{row}")
                cell_style(ws, row, 2, val, bg_c, fg_c, bold=True)
            ws.row_dimensions[row].height = 15; row += 1
        row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── BOTÓN EXPORTAR ────────────────────────────────────────────────────────────
st.divider()
st.subheader("📥 Exportar")
nombre_archivo = f"Inspeccion_{departamento or 'depto'}_{fecha}.xlsx"
st.download_button(
    "📊 Descargar Excel con formato",
    data=exportar_excel(),
    file_name=nombre_archivo,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
